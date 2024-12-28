import os
import csv
import json
import openpyxl
import sqlite3


def check_csv_file(csv_file_path):
    """检查是否存在指定的CSV文件，如果不存在则创建该文件。

        Args:
            csv_file_path (str): CSV文件路径。

        Returns:
            None
    """
    # 如果文件不存在则创建新文件
    if not os.path.exists(csv_file_path):
        with open(csv_file_path, 'w', newline='') as csvfile:
            pass


def count_images_csv(dataset_path, csv_file_path):
    """统计五种文字类型中每个汉字的图像数据量以及每个汉字对应的全部图像数据量，并将结果写入到CSV文件中。

    Args:
        dataset_path (str): 数据集文件夹路径。
        csv_file_path (str): CSV文件路径。

    Returns:
        None
    """
    # 标签转换
    json_path = './GF_class_indices.json'

    with open(json_path, "r") as f:
        class_indict = json.load(f)

    num_per_class = []
    # 定义汉字计数器和字体计数器
    check_csv_file(csv_file_path)
    type_counts = {}
    char_counts = {}
    # type_counts = { '楷书': {}, '行书': {}, '草书': {}, '隶书': {}, '篆书': {} }

    # 遍历所有子文件夹
    for type_dir in os.listdir(dataset_path):
        # 忽略mac系统中的隐藏文件夹
        if type_dir.startswith('.'):
            continue
        type_counts[type_dir] = {}
        type_path = os.path.join(dataset_path, type_dir)
        if not os.path.isdir(type_path):
            continue

        # 遍历当前类别的所有汉字文件夹
        for char_dir in os.listdir(type_path):
            # 忽略mac系统中的隐藏文件
            if char_dir.startswith('.'):
                continue
            char_path = os.path.join(type_path, char_dir)
            if not os.path.isdir(char_path):
                continue

            # 统计当前汉字文件夹中的图像数量
            image_count = len(os.listdir(char_path))

            # 更新计数器
            char_counts[char_dir] = char_counts.get(char_dir, 0) + image_count
            type_counts[type_dir][char_dir] = image_count

    # 将结果按照降序排列
    sorted_char_counts = sorted(char_counts.items(), key=lambda x: x[1], reverse=True)

    for i in range(len(class_indict)):
        hanzi = class_indict[str(i)]
        sampers = char_counts.get(hanzi, 1)
        num_per_class.append(sampers)

    # 将结果写入CSV文件中
    with open(csv_file_path, 'w', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(['汉字', '全部图像数量'] + list(type_counts.keys()))
        for char_dir, all_count in sorted_char_counts:
            row = [char_dir, all_count]
            for type_dir in type_counts.keys():
                row.append(type_counts[type_dir].get(char_dir, 0))
            writer.writerow(row)

    return num_per_class


def check_excel_file(excel_file_path):
    """检查是否存在指定的Excel文件，如果不存在则创建该文件。

        Args:
            excel_file_path (str): Excel文件路径。

        Returns:
            openpyxl.workbook.workbook.Workbook: 创建或打开的Excel工作簿。
    """
    # 如果文件存在则打开，否则创建新文件
    if os.path.exists(excel_file_path):
        wb = openpyxl.load_workbook(excel_file_path)
    else:
        wb = openpyxl.Workbook()
        wb.save(excel_file_path)

    return wb


def count_images(dataset_path, excel_file_path):
    """统计五种文字类型中每个汉字的图像数据量以及每个汉字对应的全部图像数据量，并将结果写入到Excel文件中。

    Args:
        dataset_path (str): 数据集文件夹路径。
        excel_file_path (str): Excel文件路径。

    Returns:
        None
    """
    # 标签转换
    json_path = './GF_class_indices.json'

    with open(json_path, "r") as f:
        class_indict = json.load(f)

    num_per_class = []

    # 定义汉字计数器和字体计数器
    check_excel_file(excel_file_path)
    type_counts = {}
    char_counts = {}
    # type_counts = { '楷书': {}, '行书': {}, '草书': {}, '隶书': {}, '篆书': {} }

    # 遍历所有子文件夹
    for type_dir in os.listdir(dataset_path):
        # 忽略mac系统中的隐藏文件夹
        if type_dir.startswith('.'):
            continue
        type_counts[type_dir] = {}
        type_path = os.path.join(dataset_path, type_dir)
        if not os.path.isdir(type_path):
            continue

        # 遍历当前类别的所有汉字文件夹
        for char_dir in os.listdir(type_path):
            # 忽略mac系统中的隐藏文件
            if char_dir.startswith('.'):
                continue
            char_path = os.path.join(type_path, char_dir)
            if not os.path.isdir(char_path):
                continue

            # 统计当前汉字文件夹中的图像数量
            image_count = len(os.listdir(char_path))

            # 更新计数器
            char_counts[char_dir] = char_counts.get(char_dir, 0) + image_count
            type_counts[type_dir][char_dir] = image_count

    # 将结果按照降序排列
    sorted_char_counts = sorted(char_counts.items(), key=lambda x: x[1], reverse=True)

    for i in range(len(class_indict)):
        hanzi = class_indict[str(i)]
        sampers = char_counts.get(hanzi, 1)
        num_per_class.append(sampers)

    # 将结果写入Excel文件中
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'image_counts'
    sheet.cell(row=1, column=1, value='汉字')
    sheet.cell(row=1, column=2, value='全部图像数量')
    for i, type_dir in enumerate(type_counts.keys()):
        sheet.cell(row=1, column=i+3, value=type_dir)
    row = 2
    for char_dir, all_count in sorted_char_counts:
        sheet.cell(row=row, column=1, value=char_dir)
        sheet.cell(row=row, column=2, value=all_count)
        for i, type_dir in enumerate(type_counts.keys()):
            sheet.cell(row=row, column=i+3, value=type_counts[type_dir].get(char_dir, 0))
        row += 1

    wb.save(excel_file_path)

    return num_per_class


def read_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data_dict = {}
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        key, value = row[0], row[1]
        if key is not None and value is not None:
            data_dict[key] = value
    return data_dict

def read_labels(level, label_in_level):
    conn = sqlite3.connect(r'D:\wangzhaojiang\FLENet\experiments\Ensemble_learning\labels.db')
    cur = conn.cursor()

    sql = 'SELECT hanzi FROM Labels WHERE level=? AND label_in_level=?;'
    cur.execute(sql, (level, label_in_level))

    resultList = [r[0] for r in cur.fetchall()]

    conn.close()

    return resultList

def get_samples_lists(group_nums = 4,cls_per_group = [2000,2000,2000,2105]):
    data_dict = read_excel(r"D:\wangzhaojiang\FLENet\experiments\Ensemble_learning\datasets_train.xlsx")
    samples_lists = []
    for g in range(group_nums):
        g_lists=[]
        for c in range(cls_per_group[g]):
            hanzi = read_labels(g,c)[0]
            samples = data_dict.get(hanzi, 1)
            g_lists.append(samples)
        samples_lists.append(g_lists)
    
    g2 = sum(samples_lists[1])
    g3 = sum(samples_lists[2])
    g4 = sum(samples_lists[3])
    samples_lists[0].append(g2+g3+g4)
    samples_lists[1].append(g3+g4)
    samples_lists[2].append(g4)
    samples_lists[3].append(1)
    # print(samples_lists)
    return samples_lists


#samples_lists = get_samples_lists()
#print(samples_lists[0])

if __name__ == '__main__':
    dataset_path = r"D:\wangzhaojiang\书法真迹\ygsf_zj\以观书法"
    excel_file_path = "./ygsf.xlsx"
    count_images(dataset_path, excel_file_path)
