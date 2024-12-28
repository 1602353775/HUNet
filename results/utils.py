import os 
import json
import sys
import pickle
import random
import torch
import time
import onnxruntime
import torch.nn.functional as F
from typing import Union,Tuple
from tqdm import tqdm
import matplotlib.pyplot as plt
import pandas as pd
from sampler import balance_sampler
import csv
import numpy as np
from openpyxl import Workbook, load_workbook
from loss_fn.CB_loss import CB_loss
from loss_fn.LDAM_loss import LDAMLoss,LDAMLoss_GPU
from loss_fn.IB_loss import IBLoss
from loss_fn.GCL_loss import GCLLoss
from loss_fn.FeaBal_loss import FeaBalLoss
from loss_fn.CenterLoss import CenterLoss
from loss_fn.face_loss import AngularPenaltySMLoss
import xlsxwriter

IMG_EXTENSIONS = (".jpg", ".jpeg", ".png", ".ppm", ".bmp", ".pgm", ".tif", ".tiff", ".webp")

def has_file_allowed_extension(filename: str, extensions: Union[str, Tuple[str, ...]]) -> bool:
    """Checks if a file is an allowed extension.

    Args:
        filename (string): path to a file
        extensions (tuple of strings): extensions to consider (lowercase)

    Returns:
        bool: True if the filename ends with one of given extensions
    """
    return filename.lower().endswith(extensions if isinstance(extensions, str) else tuple(extensions))

def is_image_file(filename: str) -> bool:
    """Checks if a file is an allowed image extension.

    Args:
        filename (string): path to a file

    Returns:
        bool: True if the filename ends with a known image extension
    """
    return has_file_allowed_extension(filename, IMG_EXTENSIONS)

# 读取数据集
def read_data_1(train_dir,test_dir):
    train_root = train_dir
    print(train_root)
    test_root = test_dir
    print(test_root)
    # 标签转换
    json_path = './GF_class_indices.json'

    with open(json_path, "r") as f:
        class_indict = json.load(f)

    class_indict = {value: key for key, value in class_indict.items()}

    train_fonts = [font for font in os.listdir(train_root) if os.path.isdir(os.path.join(train_root,font))]
    test_fonts = [font for font in os.listdir(test_root) if os.path.isdir(os.path.join(test_root,font))]
    
    train_images_path = []   # 存储训练集的所有图片路径
    train_images_label = []  # 存储训练集图片对应索引信息
    test_images_path = []     # 存储验证集的所有图片路径
    test_images_label = []    # 存储验证集图片对应索引信息
    
    # 遍历每个文件夹下的文件
    for font in train_fonts:
        font_dir = os.path.join(train_root,font)
        for subfolder in os.listdir(font_dir):
            subfolder_path = os.path.join(font_dir, subfolder)
            if subfolder_path.startswith('.'):
                continue
            if os.path.isdir(subfolder_path):
                if subfolder in class_indict:
                    for image_name in os.listdir(subfolder_path):
                        image_path = os.path.join(subfolder_path, image_name)
                        if image_path.startswith('.'):
                            continue
                        train_images_path.append(image_path)
                        # 标签转换
                        train_images_label.append(int(class_indict[subfolder]))
        
    for font in test_fonts:
        font_dir = os.path.join(test_root,font)
        for subfolder in os.listdir(font_dir):
            subfolder_path = os.path.join(font_dir, subfolder)
            if subfolder_path.startswith('.'):
                continue
            if os.path.isdir(subfolder_path):
                if subfolder in class_indict:
                    for image_name in os.listdir(subfolder_path):
                        image_path = os.path.join(subfolder_path, image_name)
                        if image_path.startswith('.'):
                            continue
                        test_images_path.append(image_path)
                        # 标签转换
                        test_images_label.append(int(class_indict[subfolder]))



    print("{} images were found in the dataset.".format(len(train_images_path)+len(test_images_path)))
    print("{} images for training.".format(len(train_images_path)))
    print("{} images for validation.".format(len(test_images_path)))

    return train_images_path, train_images_label, test_images_path, test_images_label

def read_data_2(dataset_path):
    # 获得train和test数据集的路径
    train_dir = os.path.join(dataset_path, "train")
    test_dir = os.path.join(dataset_path, "test")

    # 标签转换
    json_path = './GF_class_indices.jsonclass_indices.json'

    with open(json_path, "r") as f:
        class_indict = json.load(f)

    class_indict = {value: key for key, value in class_indict.items()}

    # 初始化数组
    train_images = []
    train_labels = []
    test_images = []
    test_labels = []


    # 遍历train子文件夹并获取其下的图像路径以及对应的标签值
    for subfolder in os.listdir(train_dir):
        subfolder_path = os.path.join(train_dir, subfolder)
        if subfolder_path.startswith('.'):
            continue
        if os.path.isdir(subfolder_path):
            for image_name in os.listdir(subfolder_path):
                image_path = os.path.join(subfolder_path, image_name)
                if image_path.startswith('.'):
                    continue
                train_images.append(image_path)
                # 标签转换
                train_labels.append(int(class_indict[subfolder]))

    # 遍历test子文件夹并获取其下的图像路径以及对应的标签值
    for subfolder in os.listdir(test_dir):
        subfolder_path = os.path.join(test_dir, subfolder)
        if subfolder_path.startswith('.'):
            continue
        if os.path.isdir(subfolder_path):
            for image_name in os.listdir(subfolder_path):
                image_path = os.path.join(subfolder_path, image_name)
                if image_path.startswith('.'):
                    continue
                test_images.append(image_path)
                # 标签转换
                test_labels.append(int(class_indict[subfolder]))
                # print(type(class_indict[subfolder]))

    print("{} images were found in the dataset.".format(len(train_images) + len(test_images)))
    print("{} images for training.".format(len(train_images)))
    print("{} images for validation.".format(len(test_images)))

    return train_images, train_labels, test_images, test_labels


    json_path = './GF_class_indices.json'

    with open(json_path, "r") as f:
        class_indict = json.load(f)

    class_indict = {value: key for key, value in class_indict.items()}

    # 存储每个汉字对应的不同字体的图片数量和全部图片数量总和
    data = {}

    # 存储数据集中所有图片的路径和对应的标签
    img_paths, img_labels = [], []

    # 遍历根目录下的每个子目录
    for root, dirs, files in os.walk(root_dir):
        # 当前目录为书法字体所在目录，提取出汉字和书法字体名称
        if len(dirs) > 0:
            char = os.path.basename(root)
            for font_dir in dirs:
                font_path = os.path.join(root, font_dir)

                # 统计当前书法字体所包含的图片数量
                img_count = sum([len(files) for _, _, files in os.walk(font_path)])

                # 更新 data 字典
                if char not in data:
                    data[char] = {}
                data[char][font_dir] = img_count

        # 当前目录为书法字图像所在目录，记录路径和标签
        else:
            img_labels.extend(len(files) * [os.path.basename(root)])
            img_paths.extend([os.path.join(root, file) for file in files])

    # 计算每个汉字的全部数据量
    data_sum = {char: sum(data[char].values()) for char in data}

    # 按照每个字的全部数据量对结果进行降序排列
    sorted_data = sorted(data.items(), key=lambda x: data_sum[x[0]], reverse=True)

    # 将统计结果写入csv文件中
    with open(csv_file, 'w', encoding='utf-8', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(['Char'] + list(data[sorted_data[0][0]].keys()) + ['Sum'])
        for char, font_data in sorted_data:
            row = [char] + [font_data.get(font, 0) for font in data[sorted_data[0][0]]] + [data_sum[char]]
            writer.writerow(row)

    return img_paths, img_labels

def read_sfzd (data_dir):
    """
    Count the number of images for each character and font type in the dataset,
    write the results into a CSV file, and return two lists of all image paths
    and corresponding labels.
    
    Args:
        data_dir: str. The root directory of the dataset.
        
    Returns:
        A tuple containing two lists. The first list contains all image paths,
        and the second list contains corresponding labels.
    """

    json_path = './GF_class_indices.json'

    with open(json_path, "r") as f:
        class_indict = json.load(f)

    class_indict = {value: key for key, value in class_indict.items()}
    # Initialize a dictionary to store the count of images for each character and font type
    chars_and_fonts_count = {}
    
    # Loop over all subdirectories under `data_dir`
    img_paths = []
    labels = []
    for dirpath, _, filenames in os.walk(data_dir):
        if len(filenames) == 0:
            continue
        
        # Extract character and font type from the path
        dir_tokens = dirpath.split("\\")
        char = dir_tokens[-2]
        font = dir_tokens[-1]
        
        # If this is the first time seeing this character, initialize its counter
        if char not in chars_and_fonts_count:
            chars_and_fonts_count[char] = {font: 0}
            
        # Update the counts for this character and font type
        if font not in chars_and_fonts_count[char]:
            chars_and_fonts_count[char][font] = 0
        chars_and_fonts_count[char][font] += len(filenames)
        
        # Add image paths and labels to the lists
        img_paths.extend([os.path.join(dirpath, fname) for fname in filenames])
        labels.extend([char]*len(filenames))
    
    # Write the counts to a CSV file
    with open("sfzd.csv", "w", newline="") as f:
        writer = csv.writer(f)
        
        # Write the header (font types)
        header_row = ["Char/Font"]
        for font in sorted(list(chars_and_fonts_count.values())[0].keys()):
            header_row.append(font)
        header_row.append("Total")
        writer.writerow(header_row)
        
        # Write each row of the table
        for char, counts_dict in chars_and_fonts_count.items():
            row_values = [char]
            total_count = 0
            for count in counts_dict.values():
                row_values.append(count)
                total_count += count
            row_values.append(total_count)
            writer.writerow(row_values)
            
    return img_paths, labels

def get_image_paths(folder_path):
    image_paths = []
    for root, _, files in os.walk(folder_path):
        # 遍历folder_path及其子文件夹下的所有文件
        for file_name in files:
            # 判断文件是否为JPEG或PNG格式的图片
            if file_name.lower().endswith(('.jpg', '.jpeg', '.png')):
                # 将图片文件的完整路径加入到image_paths列表中
                file_path = os.path.join(root, file_name)
                image_paths.append(file_path)
    return image_paths

def read_ygsf_zj(root_dir):
    json_path = './GF_class_indices.json'

    with open(json_path, "r") as f:
        class_indict = json.load(f)

    class_indict = {value: key for key, value in class_indict.items()}
    train_images_path = []   # 存储训练集的所有图片路径
    train_images_label = []  # 存储训练集图片对应索引信息
    fonts = [font for font in os.listdir(root_dir) if os.path.isdir(os.path.join(root_dir,font))]
    for font in fonts:
        font_dir = os.path.join(root_dir,font)
        for subfolder in os.listdir(font_dir):
            subfolder_path = os.path.join(font_dir, subfolder)
            if subfolder_path.startswith('.'):
                continue
            if os.path.isdir(subfolder_path):
                images = get_image_paths(subfolder_path)
                for image_path in images:
                    if image_path.startswith('.'):
                        continue
                    train_images_path.append(image_path)
                    # 标签转换
                    train_images_label.append(int(class_indict[subfolder]))
    return train_images_path,train_images_label

def save_wrong_predictions_to_excel(image_paths, true_labels, predicted_labels, excel_path, class_indices=None, header_added=False):
    # 如果 Excel 文件不存在，则创建一个新的工作簿
    if not os.path.exists(excel_path):
        workbook = Workbook()
        worksheet = workbook.active
        # 如果还没有添加列标题，则添加列标题
        if not header_added:
            worksheet.append(["Image Path", "True Label", "Predicted Label"])
    # 如果 Excel 文件已经存在，则打开它并获取活动工作表
    else:
        workbook = load_workbook(filename=excel_path)
        worksheet = workbook.active

    # 遍历所有图像，将错误预测的图像路径、真实标签和预测标签保存到工作表中
    for i in range(len(image_paths)):
        if true_labels[i] != predicted_labels[i]:
            worksheet.append([image_paths[i], class_indices[str(true_labels[i].item())], class_indices[str(predicted_labels[i].item())]])

    # 将工作簿保存到指定的文件路径中
    workbook.save(excel_path)

def save_wrong_predictions_to_csv(image_paths, true_labels, predicted_labels, csv_path, class_indices=None, header_added=False):
    # 写入 CSV 文件
    # 如果 CSV 文件不存在，则创建一个新文件并写入列标题（如果还没有添加）
    if not os.path.exists(csv_path):
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if not header_added:
                writer.writerow(["Image Path", "True Label", "Predicted Label"])
    # 如果 CSV 文件已经存在，则以追加模式打开它（不需要再次添加列标题）
    else:
        with open(csv_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

    # 创建一个列表用于保存错误预测的结果
    wrong_predictions = []

    # 遍历所有图像，将错误预测的图像路径、真实标签和预测标签保存到列表中
    for i in range(len(image_paths)):
        if true_labels[i] != predicted_labels[i]:
            wrong_predictions.append([image_paths[i], class_indices[str(true_labels[i].item())], class_indices[str(predicted_labels[i].item())]])

    # 将列表中的错误预测结果写入 CSV 文件
    with open(csv_path, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(wrong_predictions)

def compare_labels(image_paths, true_labels, predicted_labels):
    incorrect_indices = [i for i in range(len(true_labels)) if true_labels[i] != predicted_labels[i]]
    incorrect_image_paths = [image_paths[i] for i in incorrect_indices]
    incorrect_true_labels = [true_labels[i] for i in incorrect_indices]
    incorrect_predicted_labels = [predicted_labels[i] for i in incorrect_indices]
    return incorrect_image_paths, incorrect_true_labels, incorrect_predicted_labels

def write_lists_to_excel(image_paths, true_labels, predicted_labels, file_path,class_indices=None):
    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet()

    # 写入表头
    worksheet.write(0, 0, '图片路径')
    worksheet.write(0, 1, '图片标签')
    worksheet.write(0, 2, '预测标签')

    # 写入数据
    for i in range(len(image_paths)):
        worksheet.write(i+1, 0, image_paths[i])
        worksheet.write(i+1, 1, class_indices[str(true_labels[i].item())])
        worksheet.write(i+1, 2, class_indices[str(predicted_labels[i].item())])

    workbook.close()

def write_to_xlsx(tags, correct_nums, total_nums,  file_path):
    # 计算准确率并移动到 CPU
    accuracies = (correct_nums / total_nums) * 100.0
    accuracies_rounded = torch.round(accuracies * 10000) / 10000
    accuracies_rounded_cpu = accuracies_rounded.cpu().detach().numpy()

    # 获取汉字标签列表
    hanzi_tags = [tags[str(i)] for i in tags]

    # 转换为 Pandas 数据框格式
    data_list = [hanzi_tags, accuracies_rounded_cpu, correct_nums.cpu().detach().numpy(),
                 total_nums.cpu().detach().numpy()]
    table = pd.DataFrame(data_list).transpose()

    # 添加表头
    table.columns = ['tag', 'Accuracy', 'Correct Num', 'Total Num']
    table = table.sort_values('Accuracy', ascending=False)

    # 写入 xlsx 文件
    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
        table.to_excel(writer, index=False)

    print(f'The accuracy statistics for each category have been saved.Saved results in {file_path} successfully!')

def write_to_csv(tags, correct_nums, total_nums,  file_path):
    # 计算准确率并移动到 CPU
    accuracies = (correct_nums / total_nums) * 100.0
    accuracies_rounded = torch.round(accuracies * 10000) / 10000
    accuracies_rounded_cpu = accuracies_rounded.cpu().detach().numpy()

    # 获取汉字标签列表
    hanzi_tags = [tags[str(i)] for i in tags]

    # 转换为 Pandas 数据框格式
    data_list = [hanzi_tags, accuracies_rounded_cpu, correct_nums.cpu().detach().numpy(),
                 total_nums.cpu().detach().numpy()]
    table = pd.DataFrame(data_list).transpose()

    # 添加表头
    table.columns = ['tag', 'Accuracy', 'Correct Num', 'Total Num']
    table = table.sort_values('Accuracy', ascending=False)

    # 写入 CSV 文件
    table.to_csv(file_path, index=False, encoding='utf-8')

    print(f'The accuracy statistics for each category have been saved.Saved results in {file_path} successfully!')

def read_split_data(root: str, val_rate: float = 0.2):
    random.seed(0)  # 保证随机结果可复现 其中np.random.seed(0)的作用是使得随机数据可预测，当我们设置相同的seed，每次生成的随机数相同。
    assert os.path.exists(root), "dataset root: {} does not exist.".format(root)

    fonts = [font for font in os.listdir(root) if os.path.isdir(os.path.join(root, font))]
    HanZi_class = []
    for font in fonts:
        for cla in os.listdir(os.path.join(root, font)):
            if os.path.isdir(os.path.join(root, font, cla)) and cla not in HanZi_class:
                HanZi_class.append(cla)
    HanZi_class.sort()
    class_indices = dict((k, v) for v, k in enumerate(HanZi_class))

    json_str = json.dumps(dict((val, key) for key, val in class_indices.items()),
                          indent=4)  # json.dumps将一个Python数据结构转换为JSON   indent:参数根据数据格式缩进显示，读起来更加清晰。
    with open('class_indices.json', 'w') as json_file:
        json_file.write(json_str)

    train_images_path = []  # 存储训练集的所有图片路径
    train_images_label = []  # 存储训练集图片对应索引信息
    val_images_path = []  # 存储验证集的所有图片路径
    val_images_label = []  # 存储验证集图片对应索引信息
    every_class_num = []  # 存储每个类别的样本总数

    # 遍历每个文件夹下的文件
    count = 0
    for font in fonts:
        for cla in HanZi_class:
            cla_path = os.path.join(root, font, cla)
            if (os.path.exists(cla_path)):
                images = [os.path.join(root, font, cla, i) for i in os.listdir(cla_path) if
                          is_image_file(os.path.join(cla_path, i))]
                count += len(images)
                image_class = class_indices[cla]
                val_path = random.sample(images, k=int(len(images) * val_rate))
                for img_path in images:
                    if img_path in val_path:  # 如果该路径在采样的验证集样本中则存入验证集
                        val_images_path.append(img_path)
                        val_images_label.append(image_class)
                    else:  # 否则存入训练集
                        train_images_path.append(img_path)
                        train_images_label.append(image_class)

    sampler_train_images_path, sampler_train_images_label = balance_sampler(train_images_path, train_images_label)
    print("{} images were found in the dataset.".format(len(train_images_path) + len(val_images_path)))
    print("{} images for training.".format(len(train_images_path)))
    print("after sampled {} images for training.".format(len(sampler_train_images_path)))
    print("{} images for validation.".format(len(val_images_path)))

    plot_image = False
    if plot_image:
        # 绘制每种类别个数柱状图
        plt.bar(range(len(HanZi_class)), every_class_num, align='center')
        # 将横坐标0,1,2,3,4替换为相应的类别名称
        plt.xticks(range(len(HanZi_class)), HanZi_class)
        # 在柱状图上添加数值标签
        for i, v in enumerate(every_class_num):
            plt.text(x=i, y=v + 5, s=str(v), ha='center')
        # 设置x坐标
        plt.xlabel('image class')
        # 设置y坐标
        plt.ylabel('number of images')
        # 设置柱状图的标题
        plt.title('flower class distribution')
        plt.show()

    return train_images_path, train_images_label, val_images_path, val_images_label, sampler_train_images_path, sampler_train_images_label

def read_split_data(root: str, val_rate: float = 0.2):
    random.seed(0)    # 保证随机结果可复现 其中np.random.seed(0)的作用是使得随机数据可预测，当我们设置相同的seed，每次生成的随机数相同。
    assert os.path.exists(root), "dataset root: {} does not exist.".format(root)

    fonts = [font for font in os.listdir(root) if os.path.isdir(os.path.join(root,font))]
    HanZi_class = []
    for font in fonts:
        for cla in os.listdir(os.path.join(root,font)):
            if os.path.isdir(os.path.join(root,font,cla)) and cla not in HanZi_class:
                HanZi_class.append(cla)
    HanZi_class.sort()
    class_indices = dict((k, v) for v, k in enumerate(HanZi_class))
    
    json_str = json.dumps(dict((val, key) for key, val in class_indices.items()), indent=4)   # json.dumps将一个Python数据结构转换为JSON   indent:参数根据数据格式缩进显示，读起来更加清晰。
    with open('class_indices.json', 'w') as json_file:
        json_file.write(json_str)

    train_images_path = []   # 存储训练集的所有图片路径
    train_images_label = []  # 存储训练集图片对应索引信息
    val_images_path = []     # 存储验证集的所有图片路径
    val_images_label = []    # 存储验证集图片对应索引信息
    every_class_num = []     # 存储每个类别的样本总数

    
    # 遍历每个文件夹下的文件
    count  = 0
    for font in fonts:
        for cla in HanZi_class:
            cla_path = os.path.join(root,font,cla)
            if (os.path.exists(cla_path)):
                images = [os.path.join(root,font,cla, i) for i in os.listdir(cla_path) if is_image_file(os.path.join(cla_path,i))]
                count+=len(images)
                image_class = class_indices[cla]
                val_path = random.sample(images, k=int(len(images) * val_rate)) 
                for img_path in images:
                    if img_path in val_path:              # 如果该路径在采样的验证集样本中则存入验证集
                        val_images_path.append(img_path)
                        val_images_label.append(image_class)
                    else:                                 # 否则存入训练集
                        train_images_path.append(img_path)
                        train_images_label.append(image_class)

    sampler_train_images_path , sampler_train_images_label = balance_sampler(train_images_path,train_images_label)
    print("{} images were found in the dataset.".format(len(train_images_path)+len(val_images_path)))
    print("{} images for training.".format(len(train_images_path)))
    print("after sampled {} images for training.".format(len(sampler_train_images_path)))
    print("{} images for validation.".format(len(val_images_path)))

    plot_image = False
    if plot_image:
        # 绘制每种类别个数柱状图
        plt.bar(range(len(HanZi_class)), every_class_num, align='center')
        # 将横坐标0,1,2,3,4替换为相应的类别名称
        plt.xticks(range(len(HanZi_class)), HanZi_class)
        # 在柱状图上添加数值标签
        for i, v in enumerate(every_class_num):
            plt.text(x=i, y=v + 5, s=str(v), ha='center')
        # 设置x坐标
        plt.xlabel('image class')
        # 设置y坐标
        plt.ylabel('number of images')
        # 设置柱状图的标题
        plt.title('flower class distribution')
        plt.show()
    
    return train_images_path, train_images_label, val_images_path, val_images_label, sampler_train_images_path , sampler_train_images_label

def plot_data_loader_image(data_loader):
    batch_size = data_loader.batch_size
    plot_num = min(batch_size, 4)

    json_path = './class_indices.json'
    assert os.path.exists(json_path), json_path + " does not exist."
    json_file = open(json_path, 'r')
    class_indices = json.load(json_file)

    for data in data_loader:
        images, labels = data
        for i in range(plot_num):
            # [C, H, W] -> [H, W, C]
            img = images[i].numpy().transpose(1, 2, 0)
            # 反Normalize操作
            img = (img * [0.229, 0.224, 0.225] + [0.485, 0.456, 0.406]) * 255
            label = labels[i].item()
            plt.subplot(1, plot_num, i+1)
            plt.xlabel(class_indices[str(label)])
            plt.xticks([])  # 去掉x轴的刻度
            plt.yticks([])  # 去掉y轴的刻度
            plt.imshow(img.astype('uint8'))
        plt.show()

def write_pickle(list_info: list, file_name: str):
    with open(file_name, 'wb') as f:
        pickle.dump(list_info, f)

def read_pickle(file_name: str) -> list:
    with open(file_name, 'rb') as f:
        info_list = pickle.load(f)
        return info_list

def crossEntropy(softmax, logit, label, weight):
    target = F.one_hot(label,20)
    loss = - (weight * (target * torch.log(softmax(logit)+1e-7)).sum(dim=1)).sum()
    return loss

def weighted_loss(predictions,targets, num_classes):
    # 计算每个类别的数量
    label_counts = torch.bincount(targets, minlength=num_classes).float()

    # 计算每个类别的权重
    class_weights = 1 / label_counts
    class_weights /= class_weights.sum()

    # 计算加权交叉熵损失函数
    loss = F.cross_entropy(predictions, targets, weight=class_weights)

    return loss

def train_one_epoch(model, optimizer, data_loader, device, epoch,samples_per_cls,loss_type,loss_fn,no_of_classes):
    model.train()
    loss_function = torch.nn.CrossEntropyLoss(label_smoothing = 0.1)
    accu_loss = torch.zeros(1).to(device)  # 累计损失
    accu_num = torch.zeros(1).to(device)   # 累计预测正确的样本数
    optimizer.zero_grad()

    sample_num = 0
    data_loader = tqdm(data_loader, file=sys.stdout)
    for step, data in enumerate(data_loader):
        images, labels, _ = data
        sample_num += images.shape[0]
        
        images = images.to(device)
        labels = labels.to(device)
        # pred ,feats= model(images)
        pred = model(images)
        pred_classes = torch.max(pred, dim=1)[1]
        accu_num += torch.eq(pred_classes, labels).sum()

        # samples_per_cls_t = torch.tensor(samples_per_cls).to(device)
        # no_of_classes = torch.tensor(no_of_classes).to(device)
        if loss_fn == 'CB':
            loss = CB_loss(labels,pred,samples_per_cls,no_of_classes,loss_type,beta=0.9,gamma=2)
        elif loss_fn == 'LDAM':
            loss_c = LDAMLoss_GPU(samples_per_cls)
            loss = loss_c(pred,labels)
        elif loss_fn == 'IB':
            loss_c = IBLoss(num_classes=no_of_classes)
            loss = loss_c(pred,labels,feats)
        elif loss_fn == 'GCL':
            loss_c = GCLLoss(samples_per_cls,noise_mul = 0.2)
            loss = loss_c(pred,labels)
        elif loss_fn == 'CE+Center':
            feat_dim = feats.size(1)
            centerLoss = CenterLoss(num_classes=no_of_classes,feat_dim=feat_dim )
            loss = loss_function(pred, labels) + 0.01 * centerLoss(feats,labels)
        elif loss_fn == 'arcface':
            feat_dim = feats.size(1)
            angularloss = AngularPenaltySMLoss(feat_dim, no_of_classes, loss_type='arcface')
            loss = angularloss(feats,labels)
        else:
            loss = loss_function(pred, labels)
        
        loss.backward()
        accu_loss += loss.detach()

        data_loader.desc = "[train epoch {}] loss: {:.5f}, acc: {:.5f}".format(epoch,
                                                                               accu_loss.item() / (step + 1),
                                                                               accu_num.item() / sample_num)

        if not torch.isfinite(loss):
            print('WARNING: non-finite loss, ending training ', loss)
            sys.exit(1)

        optimizer.step()  # 参考 https://blog.csdn.net/yangwangnndd/article/details/95622893
        # 梯度累加则实现了batchsize的变相扩大，如果accumulation_steps为8，则batchsize '变相' 扩大了8倍，是我们这种乞丐实验室解决显存受限的一个不错的trick，使用时需要注意，学习率也要适当放大。
        optimizer.zero_grad()  

    return accu_loss.item() / (step + 1), accu_num.item() / sample_num


@torch.no_grad()
def evaluate_one_epoch(model, model_name, data_loader, device, epoch, excel_path,num_classes=8105,count=False):
    json_path = './GF_class_indices.json'

    with open(json_path, "r") as f:
        class_indict = json.load(f)
    
    loss_function = torch.nn.CrossEntropyLoss()
         
    model.eval()
    class_correct = torch.zeros(num_classes, device=device)
    class_total = torch.zeros(num_classes, device=device)
 
    accu_num = torch.zeros(1).to(device)   # 累计预测正确的样本数
    accu_loss = torch.zeros(1).to(device)  # 累计损失
    
    error_paths = []
    error_labels = []
    error_predict = []

    sample_num = 0
    total_time = 0
    data_loader = tqdm(data_loader, file=sys.stdout)
    for step, data in enumerate(data_loader):
        images, labels, paths = data
        sample_num += images.shape[0]

        start_time = time.time()
        # pred,feats = model(images.to(device))
        pred = model(images.to(device))
        end_time = time.time()
        total_time += end_time - start_time
        pred_classes = torch.max(pred, dim=1)[1]
        # print(pred_classes)

        # input_array = images.numpy()
        # 创建包含输入数据的字典
        # input_data = {'input': input_array}
        # providers = ['CUDAExecutionProvider', 'CPUExecutionProvider']
        # session = onnxruntime.InferenceSession(r"D:\wangzhaojiang\FLENet\mobilenet_v3_large.onnx", providers=providers)
        # output_names = session.get_outputs()
        # output_names = [output.name for output in output_names]
        # outputs = session.run(output_names, input_data, None)
        # pred2 = np.argmax(outputs[0], axis=1)
        # print(pred2)

        
        c = (pred_classes == labels.to(device)).squeeze()
        for i in range(len(labels.to(device))):
            label = labels[i]
            class_correct[label] += c[i].item()
            class_total[label] += 1
        
        accu_num += torch.eq(pred_classes, labels.to(device)).sum()
        loss = loss_function(pred, labels.to(device))
        accu_loss += loss

        data_loader.desc = "[valid epoch {}] loss: {:.5f}, acc: {:.5f}".format(epoch,
                                                                               accu_loss.item() / (step + 1),
                                                                               accu_num.item() / sample_num)
        if count:
            incorrect_image_paths, incorrect_true_labels, incorrect_predicted_labels  = compare_labels(paths, labels.to(device), pred_classes)
            error_paths.extend(incorrect_image_paths)
            error_labels.extend(incorrect_true_labels)
            error_predict.extend(incorrect_predicted_labels)

        # if count:
        #     if step == 0:
        #         save_wrong_predictions_to_excel(paths, labels.to(device), pred_classes, excel_path, class_indict, header_added=False)
        #     else:
        #         save_wrong_predictions_to_excel(paths, labels.to(device), pred_classes, excel_path, class_indict, header_added=True)

    if count:
        write_lists_to_excel(error_paths, error_labels, error_predict, excel_path,class_indict)
        if os.path.exists("./results/acc_count/{}".format(model_name)) is False:
            os.makedirs("./results/acc_count/{}".format(model_name))

        file_name = './results/acc_count/{}/best_accuracy_results_{}.xlsx'.format(model_name,epoch)
        write_to_xlsx(class_indict, class_correct, class_total, file_name)
    
    throughput = sample_num / total_time  # 计算吞吐量（样本/秒）
    print("Throughput: {:.2f} samples/sec".format(throughput))

    return accu_loss.item() / (step + 1), accu_num.item() / sample_num


def evaluate_one_epoch_onnx(model_path,  data_loader):
    json_path = './GF_class_indices.json'

    with open(json_path, "r") as f:
        class_indict = json.load(f)
    
    sample_num = 0
    acc_num = 0
    data_loader = tqdm(data_loader, file=sys.stdout)
    for step, data in enumerate(data_loader):
        images, labels, paths = data
        sample_num += images.shape[0]

        input_array = images.numpy()
        # 创建包含输入数据的字典
        input_data = {'input': input_array}
        providers = ['CUDAExecutionProvider', 'CPUExecutionProvider']
        session = onnxruntime.InferenceSession(model_path,providers=providers)
        output_names = session.get_outputs()
        output_names = [output.name for output in output_names]
        outputs = session.run(output_names, input_data, None)
        pred = np.argmax(outputs[0], axis=1)
        print(pred)
        equal_elements = pred == labels.numpy()
        print(equal_elements)
        num_equal = np.sum(equal_elements)
        acc_num += num_equal

        data_loader.desc = "[acc: {:.5f}]".format(acc_num/sample_num)
        
    return acc_num/sample_num

