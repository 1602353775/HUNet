
import os
import sys
import json
import pickle
import random
import torch
import torch.nn.functional as F
from typing import Union,Tuple
from tqdm import tqdm
from sampler import balance_sampler
import matplotlib.pyplot as plt
import pandas as pd
import csv
import numpy as np
from openpyxl import Workbook, load_workbook


IMG_EXTENSIONS = (".jpg", ".jpeg", ".png", ".ppm", ".bmp", ".pgm", ".tif", ".tiff", ".webp")

def save_wrong_predictions_to_excel(image_paths, true_labels, predicted_labels, excel_path, class_indices=None, header_added=False):
    # 如果 Excel 文件不存在，则创建一个新的工作簿
    if not os.path.exists(excel_path):
        workbook = Workbook()
        worksheet = workbook.active
        # 如果还没有添加列标题，则添加列标题
        if not header_added:
            worksheet.append(["Image Path", "True Label", "Predicted Label"])
    # 如果 Excel 文件已经存在，则打开它并获取活动工作表
    else:
        workbook = load_workbook(filename=excel_path)
        worksheet = workbook.active

    # 遍历所有图像，将错误预测的图像路径、真实标签和预测标签保存到工作表中
    for i in range(len(image_paths)):
        if true_labels[i] != predicted_labels[i]:
            worksheet.append([image_paths[i], class_indices[str(true_labels[i].item())], class_indices[str(predicted_labels[i].item())]])

    # 将工作簿保存到指定的文件路径中
    workbook.save(excel_path)


def write_to_xlsx(tags, correct_nums, total_nums,  file_path):
    # 计算准确率并移动到 CPU
    accuracies = (correct_nums / total_nums) * 100.0
    accuracies_rounded = torch.round(accuracies * 10000) / 10000
    accuracies_rounded_cpu = accuracies_rounded.cpu().detach().numpy()

    # 获取汉字标签列表
    hanzi_tags = [tags[str(i)] for i in tags]

    # 转换为 Pandas 数据框格式
    data_list = [hanzi_tags, accuracies_rounded_cpu, correct_nums.cpu().detach().numpy(),
                 total_nums.cpu().detach().numpy()]
    table = pd.DataFrame(data_list).transpose()

    # 添加表头
    table.columns = ['Hanzi', 'Accuracy', 'Correct Num', 'Total Num']
    table = table.sort_values('Accuracy', ascending=False)

    # 写入 xlsx 文件
    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
        table.to_excel(writer, index=False)

    print(f'The accuracy statistics for each category have been saved.Saved results in {file_path} successfully!')


def has_file_allowed_extension(filename: str, extensions: Union[str, Tuple[str, ...]]) -> bool:
    """Checks if a file is an allowed extension.

    Args:
        filename (string): path to a file
        extensions (tuple of strings): extensions to consider (lowercase)

    Returns:
        bool: True if the filename ends with one of given extensions
    """
    return filename.lower().endswith(extensions if isinstance(extensions, str) else tuple(extensions))


def is_image_file(filename: str) -> bool:
    """Checks if a file is an allowed image extension.

    Args:
        filename (string): path to a file

    Returns:
        bool: True if the filename ends with a known image extension
    """
    return has_file_allowed_extension(filename, IMG_EXTENSIONS)


def load_dataset(dataset_path):
    # 获得train和test数据集的路径
    train_dir = os.path.join(dataset_path, "train")
    test_dir = os.path.join(dataset_path, "test")

    # 初始化数组
    train_images = []
    train_labels = []
    test_images = []
    test_labels = []

    # 初始化标签字典
    label_map = {}
    label_index = 0

    # 遍历train子文件夹并获取其下的图像路径以及对应的标签值
    for subfolder in os.listdir(train_dir):
        subfolder_path = os.path.join(train_dir, subfolder)
        if subfolder_path.startswith('.'):
            continue
        if os.path.isdir(subfolder_path):
            for image_name in os.listdir(subfolder_path):
                image_path = os.path.join(subfolder_path, image_name)
                if image_path.startswith('.'):
                    continue
                train_images.append(image_path)
                # 如果标签不在标签映射字典中，则添加新的标签
                if subfolder not in label_map:
                    label_map[subfolder] = label_index
                    label_index += 1
                train_labels.append(label_map[subfolder])

    # 遍历test子文件夹并获取其下的图像路径以及对应的标签值
    for subfolder in os.listdir(test_dir):
        subfolder_path = os.path.join(test_dir, subfolder)
        if os.path.isdir(subfolder_path):
            for image_name in os.listdir(subfolder_path):
                image_path = os.path.join(subfolder_path, image_name)
                test_images.append(image_path)
                # 如果标签不在标签映射字典中，则添加新的标签
                if subfolder not in label_map:
                    label_map[subfolder] = label_index
                    label_index += 1
                test_labels.append(label_map[subfolder])
    
    new_dict = {v: k for k, v in label_map.items()}
    # 将标签映射保存为json文件
    with open('class_indices.json', "w") as f:
        json.dump(new_dict, f, indent=4)
    
    print("{} images were found in the dataset.".format(len(train_images)+len(test_images)))
    print("{} images for training.".format(len(train_images)))
    print("{} images for validation.".format(len(test_images)))

    return train_images, train_labels, test_images, test_labels


def load_datasets_2(root: str):
    train_root = os.path.join(root,'train')
    print(train_root)
    test_root = os.path.join(root,'test')
    print(test_root)

    # 标签转换
    json_path = './class_indices.json'

    with open(json_path, "r") as f:
        class_indict = json.load(f)

    class_indict = {value: key for key, value in class_indict.items()}

    train_fonts = [font for font in os.listdir(train_root) if os.path.isdir(os.path.join(train_root,font))]
    # print(train_fonts)
    # HanZi_class = []
    # for font in train_fonts:
    #     for cla in os.listdir(os.path.join(train_root,font)):
    #         if os.path.isdir(os.path.join(train_root,font,cla)) and cla not in HanZi_class:
    #             HanZi_class.append(cla)
    test_fonts = [font for font in os.listdir(test_root) if os.path.isdir(os.path.join(test_root,font))]
    # # print(test_fonts)
    # for font in test_fonts:
    #     for cla in os.listdir(os.path.join(test_root,font)):
    #         if os.path.isdir(os.path.join(test_root,font,cla)) and cla not in HanZi_class:
    #             HanZi_class.append(cla)
    
    # HanZi_class.sort()
    # class_indices = dict((k, v) for v, k in enumerate(HanZi_class))
    
    # json_str = json.dumps(dict((val, key) for key, val in class_indices.items()), indent=4)   # json.dumps将一个Python数据结构转换为JSON   indent:参数根据数据格式缩进显示，读起来更加清晰。
    # with open('class_indices.json', 'w') as json_file:
    #     json_file.write(json_str)

    train_images_path = []   # 存储训练集的所有图片路径
    train_images_label = []  # 存储训练集图片对应索引信息
    test_images_path = []     # 存储验证集的所有图片路径
    test_images_label = []    # 存储验证集图片对应索引信息
    
    # 遍历每个文件夹下的文件
    for font in train_fonts:
        font_dir = os.path.join(train_root,font)
        for subfolder in os.listdir(font_dir):
            subfolder_path = os.path.join(font_dir, subfolder)
            if subfolder_path.startswith('.'):
                continue
            if os.path.isdir(subfolder_path):
                for image_name in os.listdir(subfolder_path):
                    image_path = os.path.join(subfolder_path, image_name)
                    if image_path.startswith('.'):
                        continue
                    train_images_path.append(image_path)
                    # 标签转换
                    train_images_label.append(int(class_indict[subfolder]))
                    # print(type(class_indict[subfolder]))
        
    for font in test_fonts:
        font_dir = os.path.join(test_root,font)
        for subfolder in os.listdir(font_dir):
            subfolder_path = os.path.join(font_dir, subfolder)
            if subfolder_path.startswith('.'):
                continue
            if os.path.isdir(subfolder_path):
                for image_name in os.listdir(subfolder_path):
                    image_path = os.path.join(subfolder_path, image_name)
                    if image_path.startswith('.'):
                        continue
                    test_images_path.append(image_path)
                    # 标签转换
                    test_images_label.append(int(class_indict[subfolder]))
                    # print(type(class_indict[subfolder]))



    # sampler_train_images_path , sampler_train_images_label = balance_sampler(train_images_path,train_images_label)
    print("{} images were found in the dataset.".format(len(train_images_path)+len(test_images_path)))
    print("{} images for training.".format(len(train_images_path)))
    # print("after sampled {} images for training.".format(len(sampler_train_images_path)))
    print("{} images for validation.".format(len(test_images_path)))


    return train_images_path, train_images_label, test_images_path, test_images_label

def read_split_data(root: str, val_rate: float = 0.2):
    random.seed(0)    # 保证随机结果可复现 其中np.random.seed(0)的作用是使得随机数据可预测，当我们设置相同的seed，每次生成的随机数相同。
    assert os.path.exists(root), "dataset root: {} does not exist.".format(root)

    fonts = [font for font in os.listdir(root) if os.path.isdir(os.path.join(root,font))]
    HanZi_class = []
    for font in fonts:
        for cla in os.listdir(os.path.join(root,font)):
            if os.path.isdir(os.path.join(root,font,cla)) and cla not in HanZi_class:
                HanZi_class.append(cla)
    HanZi_class.sort()
    class_indices = dict((k, v) for v, k in enumerate(HanZi_class))
    
    json_str = json.dumps(dict((val, key) for key, val in class_indices.items()), indent=4)   # json.dumps将一个Python数据结构转换为JSON   indent:参数根据数据格式缩进显示，读起来更加清晰。
    with open('class_indices.json', 'w') as json_file:
        json_file.write(json_str)

    train_images_path = []   # 存储训练集的所有图片路径
    train_images_label = []  # 存储训练集图片对应索引信息
    val_images_path = []     # 存储验证集的所有图片路径
    val_images_label = []    # 存储验证集图片对应索引信息
    every_class_num = []     # 存储每个类别的样本总数

    
    # 遍历每个文件夹下的文件
    count  = 0
    for font in fonts:
        for cla in HanZi_class:
            cla_path = os.path.join(root,font,cla)
            if (os.path.exists(cla_path)):
                images = [os.path.join(root,font,cla, i) for i in os.listdir(cla_path) if is_image_file(os.path.join(cla_path,i))]
                count+=len(images)
                image_class = class_indices[cla]
                val_path = random.sample(images, k=int(len(images) * val_rate)) 
                for img_path in images:
                    if img_path in val_path:              # 如果该路径在采样的验证集样本中则存入验证集
                        val_images_path.append(img_path)
                        val_images_label.append(image_class)
                    else:                                 # 否则存入训练集
                        train_images_path.append(img_path)
                        train_images_label.append(image_class)

    # sampler_train_images_path , sampler_train_images_label = balance_sampler(train_images_path,train_images_label)
    print("{} images were found in the dataset.".format(len(train_images_path)+len(val_images_path)))
    print("{} images for training.".format(len(train_images_path)))
    # print("after sampled {} images for training.".format(len(sampler_train_images_path)))
    print("{} images for validation.".format(len(val_images_path)))

    plot_image = False
    if plot_image:
        # 绘制每种类别个数柱状图
        plt.bar(range(len(HanZi_class)), every_class_num, align='center')
        # 将横坐标0,1,2,3,4替换为相应的类别名称
        plt.xticks(range(len(HanZi_class)), HanZi_class)
        # 在柱状图上添加数值标签
        for i, v in enumerate(every_class_num):
            plt.text(x=i, y=v + 5, s=str(v), ha='center')
        # 设置x坐标
        plt.xlabel('image class')
        # 设置y坐标
        plt.ylabel('number of images')
        # 设置柱状图的标题
        plt.title('flower class distribution')
        plt.show()
    
    return train_images_path, train_images_label, val_images_path, val_images_label

def plot_data_loader_image(data_loader):
    batch_size = data_loader.batch_size
    plot_num = min(batch_size, 4)

    json_path = './class_indices.json'
    assert os.path.exists(json_path), json_path + " does not exist."
    json_file = open(json_path, 'r')
    class_indices = json.load(json_file)

    for data in data_loader:
        images, labels = data
        for i in range(plot_num):
            # [C, H, W] -> [H, W, C]
            img = images[i].numpy().transpose(1, 2, 0)
            # 反Normalize操作
            img = (img * [0.229, 0.224, 0.225] + [0.485, 0.456, 0.406]) * 255
            label = labels[i].item()
            plt.subplot(1, plot_num, i+1)
            plt.xlabel(class_indices[str(label)])
            plt.xticks([])  # 去掉x轴的刻度
            plt.yticks([])  # 去掉y轴的刻度
            plt.imshow(img.astype('uint8'))
        plt.show()


def write_pickle(list_info: list, file_name: str):
    with open(file_name, 'wb') as f:
        pickle.dump(list_info, f)


def read_pickle(file_name: str) -> list:
    with open(file_name, 'rb') as f:
        info_list = pickle.load(f)
        return info_list


def crossEntropy(softmax, logit, label, weight):
    target = F.one_hot(label,20)
    loss = - (weight * (target * torch.log(softmax(logit)+1e-7)).sum(dim=1)).sum()
    return loss


def weighted_loss(predictions,targets, num_classes):
    # 计算每个类别的数量
    label_counts = torch.bincount(targets, minlength=num_classes).float()

    # 计算每个类别的权重
    class_weights = 1 / label_counts
    class_weights /= class_weights.sum()

    # 计算加权交叉熵损失函数
    loss = F.cross_entropy(predictions, targets, weight=class_weights)

    return loss


def train_one_epoch(model, optimizer, data_loader, device, epoch, num_classes):
    model.train()
    loss_function = torch.nn.CrossEntropyLoss(label_smoothing = 0.1)
    accu_loss = torch.zeros(1).to(device)  # 累计损失
    accu_num = torch.zeros(1).to(device)   # 累计预测正确的样本数
    optimizer.zero_grad()

    sample_num = 0
    data_loader = tqdm(data_loader, file=sys.stdout)
    for step, data in enumerate(data_loader):
        images, labels, paths = data
        sample_num += images.shape[0]

        pred = model(images.to(device))

        pred_classes = torch.max(pred, dim=1)[1]

        accu_num += torch.eq(pred_classes, labels.to(device)).sum()

        loss = loss_function(pred, labels.to(device))
        # loss = weighted_loss(pred,labels.to(device),num_classes)
        loss.backward()
        accu_loss += loss.detach()

        data_loader.desc = "[train epoch {}] loss: {:.5f}, acc: {:.5f}".format(epoch,
                                                                               accu_loss.item() / (step + 1),
                                                                               accu_num.item() / sample_num)

        if not torch.isfinite(loss):
            print('WARNING: non-finite loss, ending training ', loss)
            sys.exit(1)

        optimizer.step()  # 参考 https://blog.csdn.net/yangwangnndd/article/details/95622893
        # 梯度累加则实现了batchsize的变相扩大，如果accumulation_steps为8，则batchsize '变相' 扩大了8倍，是我们这种乞丐实验室解决显存受限的一个不错的trick，使用时需要注意，学习率也要适当放大。
        optimizer.zero_grad()  

    return accu_loss.item() / (step + 1), accu_num.item() / sample_num


@torch.no_grad()
def evaluate(model, model_name, data_loader, device, epoch, excel_path,num_classes=500,count=False):
    acc_count = {}
    error = []
    json_path = './class_indices.json'

    with open(json_path, "r") as f:
        class_indict = json.load(f)
    # for index in class_indict.key():
    loss_function = torch.nn.CrossEntropyLoss()
         
    model.eval()
    class_correct = torch.zeros(num_classes, device=device)
    class_total = torch.zeros(num_classes, device=device)
 
    accu_num = torch.zeros(1).to(device)   # 累计预测正确的样本数
    accu_loss = torch.zeros(1).to(device)  # 累计损失

    sample_num = 0
    data_loader = tqdm(data_loader, file=sys.stdout)
    for step, data in enumerate(data_loader):
        images, labels, paths = data
        sample_num += images.shape[0]

        pred = model(images.to(device))
        pred_classes = torch.max(pred, dim=1)[1]

        c = (pred_classes == labels.to(device)).squeeze()
        for i in range(len(labels.to(device))):
            label = labels[i]
            class_correct[label] += c[i].item()
            class_total[label] += 1
        
        accu_num += torch.eq(pred_classes, labels.to(device)).sum()
        loss = loss_function(pred, labels.to(device))
        accu_loss += loss

        data_loader.desc = "[valid epoch {}] loss: {:.5f}, acc: {:.5f}".format(epoch,
                                                                               accu_loss.item() / (step + 1),
                                                                               accu_num.item() / sample_num)


        if count:
            if step == 0:
                save_wrong_predictions_to_excel(paths, labels, pred_classes, excel_path, class_indict, header_added=False)
            else:
                save_wrong_predictions_to_excel(paths, labels, pred_classes, excel_path, class_indict, header_added=True)

    if count:
        if os.path.exists("./acc_count/{}".format(model_name)) is False:
            os.makedirs("./acc_count/{}".format(model_name))

        file_name = './acc_count/{}/best_accuracy_results_{}.xlsx'.format(model_name,epoch)
        write_to_xlsx(class_indict, class_correct, class_total, file_name)



    return accu_loss.item() / (step + 1), accu_num.item() / sample_num

def load_state_dict(model, state_dict):
    if hasattr(model, "module"):
        model.module.load_state_dict(state_dict)
    else:
        model.load_state_dict(state_dict)
    return model

def load_checkpoint(resume_loc,model,optimizer):
    device = torch.device("cuda:0" if torch.cuda.is_available() else "cpu")
    checkpoint = torch.load(resume_loc,map_location=device)
    start_epoch = checkpoint["epoch"] + 1
    best_metric = checkpoint["best_metric"]
    model = load_state_dict(model, checkpoint["model_state_dict"])
    optimizer = load_state_dict(optimizer,checkpoint["optim_state_dict"])
    return (
        model,
        optimizer,
        start_epoch,
        best_metric,
    )


